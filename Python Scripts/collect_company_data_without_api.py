import requests
from googlesearch import search
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import time

#This method will scrape the company's official website URL and HQ address directly from the Google search engine and the company's Wiki pages. 

#It works MUCH slower compared to the version that uses SerpApi, and you need to be careful about exceeding Google's rate limit. Therefore, "collect_company_data_with_api.py" is recommended as the best version for this task.

def get_URL(company_name):
    try:
        query = f"{company_name} official website"
        for result in search(query, num_results=1):
            return result
    except Exception as e:
        if "429" in str(e):
            raise Exception("Rate Limit Exceeded, Restart the Program Later")
        print(f"Error fetching URL for {company_name}: {e}")
        return "Not Found"

def get_Country(company_name):
    wikipage = None
    query = f"{company_name} wiki"
    
    try:
        for result in search(query, num_results=1):
            wikipage = result
            break
    except Exception as e:
        if "429" in str(e):
            raise Exception("Rate Limit Exceeded, Restart the Program Later")
        print(f"Error during search for {company_name}: {e}")
        return "Not Found"

    if not wikipage or not wikipage.startswith("http"):
        return "Not Found"

    try:
        response = requests.get(wikipage)
        if response.status_code == 429:
            raise Exception("Rate Limit Exceeded, Restart the Program Later")
        if response.status_code != 200:
            return "Not Found"

        soup = BeautifulSoup(response.text, 'html.parser')
        infobox = soup.find('table', {'class': 'infobox'})
        if not infobox:
            return "Not Found"

        for row in infobox.find_all('tr'):
            header = row.find('th')
            data = row.find('td')
            if header and 'Headquarters' in header.get_text() and data:
                country = data.get_text()
                return country.strip()
    except Exception as e:
        print(f"Error fetching country for {company_name}: {e}")
        raise
    return "Not Found"

try:
    book = load_workbook('Book2.xlsx')
    sheet = book.active

    for row in range(2, sheet.max_row + 1):
        company_name = sheet[f"A{row}"].value

        if not sheet[f"A{row}"].value:
            break
        
        if sheet[f"B{row}"].value and sheet[f"C{row}"].value:
            print(f"Row {row} already processed. Skipping...")
            continue

        if company_name:
            try:
                country = get_Country(company_name)
                time.sleep(2)
                url = get_URL(company_name)
                time.sleep(2)
                sheet[f"B{row}"] = country
                sheet[f"C{row}"] = url
            except Exception as e:
                if "Rate Limit Exceeded" in str(e):
                    print("Rate Limit Exceeded, Restart the Program Later")
                    break
                print(f"Error processing row {row}: {e}")
        
# Save results every 25 rows
        if row % 2 == 0:
            book.save('Book2.xlsx')
            print(f"Workbook saved after processing row {row}")

finally:
    book.save('Book2.xlsx')
    print("Workbook saved successfully.")

